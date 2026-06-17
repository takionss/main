import urllib.request
import csv
import io
import os
import json
import openpyxl

URL = "https://docs.google.com/spreadsheets/d/1RRpkarZGReZMb243IsGNQfP1Urzs5Fh24uWRbfkwVdM/export?format=xlsx"
LANGUAGES = ['ko', 'en', 'es', 'ja', 'zh-TW']
WALLPAPER_GROUPS = ['배경화면', 'Wallpaper', 'Wallpapers', 'Background', 'Backgrounds', 'Fondo de pantalla', 'Papel pintado', '壁紙', '背景', '背景图片']

def resolve_url(url_str):
    if not url_str:
        return ""
    if url_str.startswith("http://") or url_str.startswith("https://") or url_str.startswith("//"):
        return url_str
    return "/" + url_str.lstrip("/")


TRANSLATIONS = {
    'ko': {
        't_title': 'SmilesLife Portal - 일반 모드',
        't_description': '종합 웹 도구와 다양한 게임을 카테고리별로 정렬하여 바로 실행할 수 있는 일반 모드입니다.',
        't_categories': 'Categories',
        't_win_mode': '윈도우 모드로 변경',
        't_reset': '선택 모드 초기화'
    },
    'en': {
        't_title': 'SmilesLife Portal - Normal Mode',
        't_description': 'A collection of web utility tools and games sorted by category. Launch them instantly.',
        't_categories': 'Categories',
        't_win_mode': 'Switch to Windows Mode',
        't_reset': 'Reset Preferences'
    },
    'es': {
        't_title': 'SmilesLife Portal - Modo Normal',
        't_description': 'Una colección de herramientas web útiles y juegos organizados por categorías. Ejecútalos al instante.',
        't_categories': 'Categorías',
        't_win_mode': 'Cambiar al modo Windows',
        't_reset': 'Restablecer preferencias'
    },
    'ja': {
        't_title': 'SmilesLife Portal - ノー말모드',
        't_description': '便利なウェブツールや様々なゲーム를 カテゴリ別に整理し、すぐに実行できるノーマルモードです。',
        't_categories': 'カテゴリー',
        't_win_mode': 'ウィンドウモードに変更',
        't_reset': '選択モード의 초기화'
    },
    'zh-TW': {
        't_title': 'SmilesLife Portal - 一般模式',
        't_description': '將綜合網頁工具與多樣遊戲按分類整理，可立即執行的一般模式。',
        't_categories': '分類',
        't_win_mode': '切換至 Windows 模式',
        't_reset': '重置偏好設定'
    }
}

def main():
    print(f"Downloading Google Sheets Excel from {URL}...")
    temp_filename = "temp_sheets.xlsx"
    try:
        req = urllib.request.Request(
            URL, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        )
        with urllib.request.urlopen(req) as response:
            with open(temp_filename, 'wb') as f:
                f.write(response.read())
    except Exception as e:
        print(f"Error downloading Excel file: {e}")
        return

    # Load Excel file
    try:
        wb = openpyxl.load_workbook(temp_filename, data_only=True)
    except Exception as e:
        print(f"Error loading Excel workbook: {e}")
        if os.path.exists(temp_filename):
            os.remove(temp_filename)
        return

    # 1. Resolve the primary Korean (ko) sheet
    ko_sheet_name = None
    possible_ko_names = ['ko', '시트1', 'Sheet1']
    for name in possible_ko_names:
        if name in wb.sheetnames:
            ko_sheet_name = name
            break
            
    if not ko_sheet_name and len(wb.sheetnames) > 0:
        ko_sheet_name = wb.sheetnames[0]
        
    if not ko_sheet_name:
        print("Error: No sheets found in the workbook.")
        wb.close()
        os.remove(temp_filename)
        return
        
    print(f"Resolved primary (ko) sheet as: '{ko_sheet_name}'")

    def parse_sheet(sheet):
        rows_data = []
        default_icon = "images/world.png"
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx == 1:
                continue # Skip header
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
                
            row_list = list(row)
            while len(row_list) < 5:
                row_list.append('')
                
            name = str(row_list[0] or '').strip()
            group = str(row_list[1] or '').strip()
            url = str(row_list[2] or '').strip()
            icon = str(row_list[3] or '').strip()
            desc = str(row_list[4] or '').strip()
            
            # Skip rows containing spreadsheet error messages
            error_indicators = ['#value!', '#n/a', '#ref!', '#name?']
            if any(err in name.lower() or err in group.lower() or err in url.lower() or err in desc.lower() for err in error_indicators):
                continue
            
            if not group:
                group = "기타"
            if not icon:
                icon = default_icon
                
            rows_data.append({
                'name': name,
                'group': group,
                'url': url,
                'icon': icon,
                'desc': desc
            })
        return rows_data

    # Parse primary data
    ko_rows = parse_sheet(wb[ko_sheet_name])
    
    # Load HTML template
    template_path = os.path.join("templates", "index_normal_template.html")
    if not os.path.exists(template_path):
        print(f"Error: Template file {template_path} not found.")
        wb.close()
        os.remove(temp_filename)
        return
        
    with open(template_path, 'r', encoding='utf-8') as f:
        html_template = f.read()

    multilang_data = {}
    
    for lang in LANGUAGES:
        sheet_rows = None
        if lang == 'ko':
            sheet_rows = ko_rows
        elif lang in wb.sheetnames:
            print(f"Processing language sheet: {lang}")
            sheet_rows = parse_sheet(wb[lang])
        else:
            print(f"Language sheet '{lang}' not found. Falling back to primary (ko) sheet data.")
            sheet_rows = ko_rows
            
        # 1. Save language-specific CSV cache
        filepaths = [f"tools_{lang}.csv"]
        if lang == 'ko':
            filepaths.append("tools.csv")
            
        fieldnames = ['name', 'group', 'url', 'icon', 'desc']
        for filepath in filepaths:
            with open(filepath, 'w', encoding='utf-8', newline='') as out_f:
                writer = csv.DictWriter(out_f, fieldnames=fieldnames)
                writer.writeheader()
                for row in sheet_rows:
                    writer.writerow(row)
                    
        # 2. Group rows for templating
        grouped = []
        group_map = {}
        for row in sheet_rows:
            gname = row['group']
            if not row['url'] or gname in WALLPAPER_GROUPS:
                continue
                
            if gname not in group_map:
                group_obj = {'name': gname, 'items': []}
                grouped.append(group_obj)
                group_map[gname] = group_obj
            group_map[gname]['items'].append({
                'name': row['name'],
                'url': row['url'],
                'icon': row['icon'],
                'desc': row['desc']
            })
            
        multilang_data[lang] = grouped

        # 3. Generate static HTML page directly
        sidebar_menu_html = ""
        portal_content_html = ""
        
        for idx, group in enumerate(grouped):
            active_class = " active" if idx == 0 else ""
            sidebar_menu_html += f'                <li class="menu-item{active_class}" onclick="scrollToGroup(\'group-{idx}\', this)">\n'
            sidebar_menu_html += f'                    <i class="fa-solid fa-folder-open"></i> <span>{group["name"]}</span>\n'
            sidebar_menu_html += '                </li>\n'
            
            portal_content_html += f'                <section class="category-section" id="group-{idx}">\n'
            portal_content_html += f'                    <h2 class="category-title"><i class="fa-regular fa-star text-blue-500"></i> {group["name"]}</h2>\n'
            portal_content_html += '                    <div class="cards-grid">\n'
            
            for item in group["items"]:
                desc_html = f'                            <div class="item-desc">{item["desc"]}</div>\n' if item["desc"] else ""
                resolved_url = resolve_url(item["url"])
                resolved_icon = resolve_url(item["icon"])
                portal_content_html += f'                        <a class="item-card" href="{resolved_url}" target="_blank">\n'
                portal_content_html += f'                            <img src="{resolved_icon}" class="item-icon" alt="{item["name"]}">\n'
                portal_content_html += f'                            <div class="item-name">{item["name"]}</div>\n'
                portal_content_html += desc_html
                portal_content_html += '                        </a>\n'
                
            portal_content_html += '                    </div>\n'
            portal_content_html += '                </section>\n'

        # Fetch translations
        trans = TRANSLATIONS.get(lang, TRANSLATIONS['ko'])
        
        # Populate template using replace to avoid single curly brace errors in JS/CSS
        lang_html = html_template
        lang_html = lang_html.replace("{lang}", lang)
        lang_html = lang_html.replace("{t_title}", trans['t_title'])
        lang_html = lang_html.replace("{t_description}", trans['t_description'])
        lang_html = lang_html.replace("{t_categories}", trans['t_categories'])
        lang_html = lang_html.replace("{t_win_mode}", trans['t_win_mode'])
        lang_html = lang_html.replace("{t_reset}", trans['t_reset'])
        lang_html = lang_html.replace("{sidebar_menu}", sidebar_menu_html)
        lang_html = lang_html.replace("{portal_content}", portal_content_html)
        
        # Determine output file path
        if lang == 'ko':
            out_path = "index_normal.html"
        else:
            os.makedirs(lang, exist_ok=True)
            out_path = os.path.join(lang, "index_normal.html")
            
        print(f"Writing static page to {out_path}...")
        with open(out_path, 'w', encoding='utf-8') as out_h:
            out_h.write(lang_html)

    # Save to _data/tools_multilang.json
    os.makedirs("_data", exist_ok=True)
    print("Saving combined JSON to _data/tools_multilang.json...")
    with open('_data/tools_multilang.json', 'w', encoding='utf-8') as out_json:
        json.dump(multilang_data, out_json, ensure_ascii=False, indent=2)

    # Clean up temp file
    wb.close()
    if os.path.exists(temp_filename):
        os.remove(temp_filename)
    print("Success! Multi-language static HTML generation complete.")

if __name__ == "__main__":
    main()
